"""
Побудова Excel-звітів (продажі, ремонти) - те, чого не вистачало
застосунку порівняно з Telegram-ботом.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_repairs_excel(rows, parts_by_repair=None):
    """rows: список кортежів (id, photo_path, receipt_number, intake_date, completion_date, cost, payment_method).
    parts_by_repair: {repair_id: [(link, note, status), ...]} - запчастини, прив'язані до кожної квитанції."""
    parts_by_repair = parts_by_repair or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Ремонти"
    ws.append(["Номер квитанції", "Дата прийняття", "Дата видачі", "Статус", "Сума", "Оплата", "Запчастини"])
    _style_header(ws)

    total = 0.0
    for repair_id, photo_path, receipt_number, intake_date, completion_date, cost, payment_method in rows:
        status = "Видано" if completion_date else "В ремонті"
        parts = parts_by_repair.get(repair_id, [])
        parts_text = "; ".join(
            f"{p[1] or p[0]} ({p[0]})" if p[1] else p[0] for p in parts
        ) if parts else ""
        ws.append([receipt_number, intake_date, completion_date or "", status, cost or "", payment_method or "", parts_text])
        total += cost or 0

    ws.append([])
    ws.append(["Разом:", "", "", "", total, "", ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
